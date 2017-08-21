from app import app, db
from openpyxl import load_workbook
from sqlalchemy import Column, String, Integer, MetaData, Table, ForeignKey
from sqlalchemy.sql import text
from app.sales.models import Transaction, TransactionStatus
from unicodedata import normalize
from app.sales.services import create_transaction, create_merchant, get_sales_transactions_within_date_range, \
    find_merchant
from datetime import datetime
import json
import requests
from app.utils.gis_json_fields import PointToLatLngParam, PointToLatLng
from app.utils.google_api import geocode
import logging
import itertools
import re
import dateparser
from random import randint

log = logging.getLogger(__name__)

engine = db.engine
# engine = db.get_engine(app)

_punct_re = re.compile(r'[\t !"#$%&\'()*\-/<=>?@\[\\\]^_`{|},.]+')


def slugify(text, delim=u'_'):
    """Generates an slightly worse ASCII-only slug."""
    result = []
    for word in _punct_re.split(text.lower()):
        word = normalize('NFKD', word).encode('ascii', 'ignore')
        if word:
            result.append(word)
    return unicode(delim.join(result))


def clean_date(datestr, sheet):
    if sheet == 'COVERAGE':
        dateext = datestr[2:24]
        return datetime.strptime(dateext, "%m/%d/%Y-%I:%M:%S %p")
    elif sheet == 'C3S':
        return datestr
    elif sheet == 'IIDACS':
        return datestr
    elif sheet == 'LEAVES':
        return datetime.strptime(datestr, "%m/%d/%Y")
    elif sheet == 'FLEET':
        return datetime.strptime(datestr, "%m/%d/%Y")
    else:
        return dateparser.parse(datestr)


def create_table_name(name, prefix='transaction'):
    return prefix + '_' + name.lower()


address_columns = ['clinic address', 'address3', 'station address']
date_columns = ['visit_date', 'cardtransdate', 'date_time', 'leave_date', 'statementdate']

user_id_columns = ['emp id', 'employeeid', 'employee id', 'empid', 'payee id']
latlng_columns = ['latitude', 'longitude']
merchant_columns = ['mdname', 'specialty description']

except_columns = user_id_columns + latlng_columns + merchant_columns + address_columns

except_sheets = ['1SS']
# except_sheets = ['COVERAGE', '1SS', 'C3S', 'IIDACS']


def upload_fraud_data(file):
    wb = load_workbook(file, read_only=True)

    ptolatlng = PointToLatLng()
    result = {}

    # date_dict = {}
    is_table_non_transaction = False

    for sheetname in wb.get_sheet_names():
        sheet = sheetname.upper()

        if sheet in except_sheets:
            print "BYPASSING {0}. PROCEEDING TO NEXT SHEET.".format(sheet)
            continue

        print "************************"
        print sheetname
        print "************************"

        ws = wb.get_sheet_by_name(sheetname)

        rows = ws.iter_rows()

        columnrow = rows.next()
        actualcolumns = []

        tablecolumns = []
        tablecolumns_index = {}

        # data_types = {}

        table_name = create_table_name(sheetname)
        metadata = MetaData(bind=engine)

        # is_table_non_transaction = True if 'leave' in table_name.lower() else False

        for cell in columnrow:
            if cell.value is not None:
                actualcolumns.append(cell.value)

            if cell.value is None or cell.value.lower() in except_columns:
                continue

            col = slugify(cell.value)
            tablecolumns_index[str(cell.value)] = col
            tablecolumns.append(col)

        # create the table
        # if is_table_non_transaction:
        #     table = Table(table_name, metadata,
        #                   Column('id', Integer, primary_key=True),
        #                   Column('userid', Integer, ForeignKey(User.id), nullable=False),
        #                   *(Column(colname, String()) for colname in tablecolumns))
        # else:
        table = Table(table_name, metadata,
                      Column('id', Integer, primary_key=True),
                      Column('transactionid', Integer, ForeignKey(Transaction.id), nullable=False),
                      *(Column(colname, String()) for colname in tablecolumns))

        if not table.exists():
            table.create()

        values = []

        transaction_ids = []

        for row in rows:
            dct = {}
            latlng = {}
            # userid = None
            userid = 2
            merchant_data = {}
            address = None
            transaction_date = None
            merchant = None

            if len([c for c in row if c.value is not None]) == 0:
                continue

            for cell in row:
                try:
                    cls = tablecolumns_index[actualcolumns[cell.column - 1]]
                    # print "cls: {0}".format(cls)
                    if str(cls).lower() not in except_columns:
                        if cell.value is not None or cell.value != '':
                            dct[cls] = cell.value

                            if cls in date_columns:
                                # date_dict[cls] = cell.value
                                transaction_date = clean_date(cell.value, sheet)
                            # elif cls == 'station_name':
                            #     merchant_data['name'] = cell.value

                            # data_types[cls] = cell.data_type
                            # app.logger.info("{0} : {1}".format(cls, cell.value))
                        else:
                            dct[cls] = None
                except KeyError as e:
                    key = str(e).strip().lower()

                    # if len([s for s in user_id_columns if s in key]) > 0:
                    #     pass
                        # if is_table_non_transaction:
                        #     dct['userid'] = cell.value
                        # else:
                        #     userid = cell.value
                    if len([s for s in address_columns if s in key]) > 0:
                        address = cell.value.encode('utf-8')
                        merchant_data['address'] = address
                        # print "ADDRESS: {0}".format(address)
                    elif len([s for s in latlng_columns if s in key]) > 0:
                        if 'latitude' in key:
                            latlng['lat'] = cell.value
                        elif 'longitude' in key:
                            latlng['lng'] = cell.value
                    elif len([s for s in merchant_columns if s in key]) > 0:
                        if "mdname" in key:
                            merchant_data['name'] = cell.value
                        elif "specialty description" in key:
                            merchant_data['specialty'] = cell.value.upper()
                except Exception as error:
                    print "ERROR: {0}".format(error)

            # print "DATA TYPES: {0}".format(data_types)
            dct['userid'] = userid

            if userid is not None and not is_table_non_transaction:
                # save merchant
                merchantid = None

                if bool(merchant_data):
                    if 'name' in merchant_data and address is not None:
                        merchant = find_merchant(merchant_data['name'], address)

                    try:
                        if merchant is None:
                            if not bool(latlng):
                                geocode_result = geocode(merchant_data['address'])
                                latlng = geocode_result['geometry']['location']
                            merchant_data['latlng'] = latlng
                            merchant = create_merchant(merchant_data)
                            merchantid = merchant.id
                            app.logger.info("SHEET: {0} MERCHANT CREATED: {1}".format(sheet, merchant_data))
                        else:
                            merchantid = merchant.id
                            latlng = ptolatlng.format(merchant.latlng)
                            # merchant.latlng = forms_helper.parse_coordinates(latlng) if latlng is not None else merchant.latlng
                            # db.session.commit()
                    except Exception as e:
                        print "ERROR: {0}".format(e)
                        db.session.rollback()

                print "CREATING TRANSACTION: userid={0} sheet={1} latlng={2} address={3} transaction_date={4} merchantid={5}".format(
                    userid, sheet, latlng, address, transaction_date, merchantid)
                try:
                    transaction = create_transaction(userid, sheet, latlng, address, transaction_date, merchantid)
                    app.logger.info("SHEET: {0} TRANSACTION CREATED: {1}".format(sheet, transaction.to_dict()))

                    if transaction is not None and transaction.id is not None:
                        dct['transactionid'] = transaction.id
                        values.append(dct)
                        transaction_ids.append(transaction.id)
                except Exception as e:
                    print "ERROR: {0}".format(e)
                    db.session.rollback()
                    continue

            # elif is_table_non_transaction:
            #     values.append(dct)

        try:
            app.logger.info("INSERTING DATA TO {0}".format(table_name))
            # returning_column = table.c.id if is_table_non_transaction else table.c.transactionid
            returning_column = table.c.transactionid
            res = engine.execute(table.insert().values(values).returning(returning_column))
            result[sheetname.lower()] = map(lambda itm: itm[0], res)
            app.logger.info("FINISH!")
        except Exception as err:
            app.logger.info("ERROR INSERTING: {0}".format(err))

        app.logger.info("PROCEEDING NEXT SHEET")
    return result


def get_distance(origin_list, destination_list):
    pointToLatLngParam = PointToLatLngParam()

    url = 'https://maps.googleapis.com/maps/api/distancematrix/json??key=' + app.config.get('GOOGLE_MAP_API_KEY')
    url += '&origins=' + ','.join([pointToLatLngParam.format(point) for point in origin_list])
    url += '&destinations=' + ','.join([pointToLatLngParam.format(point) for point in destination_list])
    # print url

    response = requests.get(url, verify=False)

    if response.content:
        content = json.loads(response.content)
        if content['status'] == 'OK' and len(content['rows']) > 0 and len(content['rows'][0]):
            result = content['rows'][0]['elements'].pop()
            return {
                'distance': float((result['distance']['text']).split()[0]),
                'duration': float((result['duration']['text']).split()[0])
            }

    return None


def compute_transaction_travel_details(start_date, end_date, user_id):
    print "compute_transaction_travel_details startdate={0} enddate={1} userid={2}".format(start_date, end_date,
                                                                                           user_id)
    # get all the transactions within date range
    transactions = get_sales_transactions_within_date_range(start_date, end_date, user_id)

    # group transactions by day
    grouped_transactions = []
    for key, items in itertools.groupby(transactions, lambda transaction: str(transaction.transaction_date)[:10]):
        grouped_transactions.append(list(items))

    for outer_idx, listitem in enumerate(grouped_transactions):

        print "list {0} : {1}".format(outer_idx, len(listitem))

        for idx, transitm in enumerate(listitem):

            print "Transaction id={0} : {1}".format(transitm.id, transitm.transaction_date)

            latlngs = {
                'current': None,
                'previous': None,
                'next': None
            }

            transaction_dates = {
                'current': None,
                'previous': None
            }

            # refer to previous transaction (idx - 1) for previous
            # refer to next transaction (idx + 1) for next
            if len(listitem) > 1:
                if idx == 0:  # NO PREVIOUS TRANSACTION
                    # refer to start_point_latlng for previous
                    next = listitem[idx + 1]

                    latlngs['current'] = transitm.end_point_latlng
                    latlngs['next'] = next.end_point_latlng
                    latlngs['previous'] = transitm.start_point_latlng

                    transaction_dates['current'] = transitm.transaction_date
                elif idx == len(listitem) - 1:  # NO NEXT TRANSACTION
                    # do not compute next_average_travel_time and next_travel_distance
                    previous = listitem[idx - 1]

                    latlngs['current'] = transitm.end_point_latlng
                    latlngs['previous'] = previous.end_point_latlng

                    transaction_dates['current'] = transitm.transaction_date
                    transaction_dates['previous'] = previous.transaction_date
                else:
                    previous = listitem[idx - 1]
                    next = listitem[idx + 1]

                    latlngs['current'] = transitm.end_point_latlng
                    latlngs['next'] = next.end_point_latlng
                    latlngs['previous'] = previous.end_point_latlng

                    transaction_dates['current'] = transitm.transaction_date
                    transaction_dates['previous'] = previous.transaction_date

                    # print "latlngs: {0}".format(latlngs)
                    # print "transaction dates: {0}".format(transaction_dates)

            else:  # only one on list
                latlngs['current'] = transitm.end_point_latlng
                latlngs['previous'] = transitm.start_point_latlng

                transaction_dates['current'] = transitm.transaction_date

            average_travel_time_in_minutes = None
            travel_distance_in_km = None
            travel_time_in_minutes = None
            travel_time_difference = None
            next_average_travel_time_in_minutes = None
            next_travel_distace_in_km = None

            # fill average_travel_time and travel_distance_in_km by getting the previous transaction is has one if not then refer on the start_point_latlng then it in distance matrix
            # compute for average_travel_time_in_minutes and travel_distance_in_km
            if latlngs['previous'] is not None and latlngs['current'] is not None:
                l1 = [latlngs['previous']]
                l2 = [latlngs['current']]

                response = get_distance(l1, l2)

                if response is not None:
                    travel_distance_in_km = response['distance']
                    average_travel_time_in_minutes = response['duration']
                    transitm.travel_distance_in_km = travel_distance_in_km
                    transitm.average_travel_time_in_minutes = average_travel_time_in_minutes
                    print "travel_distance_in_km: {0} km average_travel_time_in_minutes: {1} mins".format(
                        travel_distance_in_km, average_travel_time_in_minutes)

            # fill next_average_travel_time_in_minutes and next_travel_distance_in_km by getting the next transaction if has one then compute it in distance matrix
            # compute for next_average_travel_time_in_km and next_travel_distance_in_km
            if latlngs['current'] is not None and latlngs['next'] is not None:
                l1 = [latlngs['current']]
                l2 = [latlngs['next']]

                response = get_distance(l1, l2)

                if response is not None:
                    next_travel_distance_in_km = response['distance']
                    next_average_travel_time_in_minutes = response['duration']
                    transitm.next_travel_distance_in_km = next_travel_distance_in_km
                    transitm.next_average_travel_time_in_minutes = next_average_travel_time_in_minutes
                    print "next_travel_distance_in_km: {0} km next_average_travel_time_in_minutes: {1} mins".format(
                        next_travel_distance_in_km, next_average_travel_time_in_minutes)

            # fill travel time based on transaction_date - previous_transaction_date
            if transaction_dates['previous'] and transaction_dates['current']:
                # compute for travel_time_in_minutes
                travel_time_result = transaction_dates['current'] - transaction_dates['previous']
                travel_time_in_minutes = round(travel_time_result.total_seconds() / 60, 2)
                transitm.travel_time_in_minutes = travel_time_in_minutes
                print "travel time in minutes: {0}".format(travel_time_in_minutes)

                if average_travel_time_in_minutes is not None:
                    scan_for_fraud_by_date(transaction_dates['current'], transaction_dates['previous'],
                                           average_travel_time_in_minutes, transitm)

            elif average_travel_time_in_minutes is not None:
                # add random additional time for the sake of test data
                transitm.travel_time_in_minutes = average_travel_time_in_minutes + randint(10, 25)

            if transitm.travel_time_in_minutes is not None:
                transitm.travel_time_difference = transitm.travel_time_in_minutes

    db.session.commit()


def scan_for_fraud_by_date(current_transaction_date, previous_transaction_date, average_travel_time_in_minutes,
                           transaction_obj):
    actual_travel_time_result = current_transaction_date - previous_transaction_date
    actual_travel_time_in_minutes = round(actual_travel_time_result.total_seconds() / 60, 2)

    if actual_travel_time_in_minutes < (average_travel_time_in_minutes - (average_travel_time_in_minutes * 0.5)) \
            or actual_travel_time_in_minutes > (
                        average_travel_time_in_minutes * 3):  # less than 50% or greater than 200% of the average travel time
        transaction_obj.status = TransactionStatus.FRAUD
        transaction_obj.remarks = 'suspicious travel time'


def scan_out_of_territory_fraud(userid):
    query = text("SELECT scan_out_of_territory_fraud({0})".format(userid)).execution_options(autocommit=True)
    print "scan out of territory fraud: userid={0}".format(userid)
    print query
    db.engine.execute(query)


def scan_on_leave_fraud(userid):
    query = text("SELECT scan_on_leave_fraud({0})".format(userid)).execution_options(autocommit=True)
    print "scan_on_leave_fraud: userid={0}".format(userid)
    print query
    db.engine.execute(query)
