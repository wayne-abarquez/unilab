from app import app, db
from .models import Branch, BranchStatus, BranchType, Product, BranchProduct, MERCHANT_SPECIALTIES, Merchant, \
    Transaction, Sellout
from app.home.models import Boundary, Territory
from sqlalchemy import select, func, desc, DATE, cast as cast_data, and_
from sqlalchemy.sql.expression import cast
from app.utils.response_transformer import to_dict
from geoalchemy2 import Geography
from app.utils import forms_helper
from random import randint
from .exceptions import BranchNotFoundError
from datetime import datetime
from openpyxl import load_workbook
import time
import requests
import json
import logging

log = logging.getLogger(__name__)


def get_branch_within_boundary(boundaryid):
    qt = select([Boundary.geometry]).select_from(Boundary).where(Boundary.id == boundaryid).alias('qt')

    stmt = select([Branch.id, Branch.type, Branch.name, Branch.latlng]) \
        .select_from(Branch) \
        .where(func.ST_DWITHIN(cast(qt.c.geometry, Geography), cast(Branch.latlng, Geography), 1)) \
        # stmt = select([Branch.id, Branch.type, Branch.name, Branch.latlng]) \
    #     .select_from(Branch) \
    #     .where(func.ST_DWITHIN(cast(qt.c.geometry, Geography), cast(Branch.latlng, Geography), 1)) \
    # .limit(500)

    result = db.engine.execute(stmt).fetchall()

    modresult = []

    for item in result:
        modresult.append({'branchid': item.id, 'boundaryid': boundaryid, 'branch': item})

    return modresult


def get_transaction_count_with_dates(dates, user_id):
    date_list = dates.split('|')

    datestr = ''
    for idx, dateitem in enumerate(date_list):
        datestr += "'" + dateitem + "'"
        if idx < len(date_list) - 1:
            datestr += ","

    query = "SELECT TO_CHAR(transaction_date, 'YYYY-MM-DD') AS transdate, COUNT(*) as ctr FROM transaction WHERE userid = {0} AND status = 'FRAUD' AND TO_CHAR(transaction_date, 'YYYY-MM-DD') IN ({1}) GROUP BY transdate".format(
        user_id, datestr)

    result = db.engine.execute(query).fetchall()

    return [{'date_param': item[0], 'count': item[1]} for item in result]


def get_branches_by_boundary(boundaryid=None):
    if boundaryid is None:
        return Branch.query.all()

    return get_branch_within_boundary(boundaryid)


def get_branches_with_limit(count=5):
    if count == 0:
        return Branch.query.all()
    else:
        return Branch.query.limit(count).all()


def get_branches_id_with_limit(count=0):
    if count == 0:  # get all
        return map(lambda idtup: idtup[0], db.session.query(Branch.id).all())
    else:
        return map(lambda idtup: idtup[0], db.session.query(Branch.id).limit(count).all())


def get_branches():
    return get_branches_by_boundary()


def get_products():
    return Product.query.all()


def get_products_by_branch(branchid):
    result = db.session.query(BranchProduct, Product) \
        .join(Product, BranchProduct.productid == Product.id) \
        .filter(BranchProduct.branchid == branchid) \
        .all()

    rp = []

    for tp in result:
        itm = to_dict(tp[0])
        itm.update({'product': to_dict(tp[1])})
        rp.append(itm)

    return rp


def create_branch(data):
    # Prepare Data
    branch = Branch.from_dict(data)
    branch.status = BranchStatus.ACTIVE
    branch.latlng = forms_helper.parse_coordinates(data['latlng'])

    # Persist
    db.session.add(branch)
    db.session.commit()

    return branch


def delete_branch(branchid):
    # Prepare Data
    branch = Branch.query.get(branchid)
    if branch is None:
        raise BranchNotFoundError("Branch id={0} not found".format(branchid))

    db.session.delete(branch)
    db.session.commit()


def create_merchant(data):
    found = Merchant.query.filter(Merchant.name == data['name'], Merchant.address == data['address']).first()

    if found is not None:
        return found

    # Prepare Data
    merchant = Merchant.from_dict(data)
    merchant.latlng = forms_helper.parse_coordinates(data['latlng'])

    if 'specialty' not in data:
        merchant.specialty = MERCHANT_SPECIALTIES[randint(0, len(MERCHANT_SPECIALTIES) - 1)]

    # Persist
    db.session.add(merchant)
    db.session.commit()

    return merchant


def get_sales_transactions():
    return Transaction.query.all()


def create_transaction(userid, endpoint_latlng, type, merchantid=None, address=None, transaction_date=datetime.now()):
    transaction = Transaction(userid=userid, type=type, transaction_date=transaction_date)
    transaction.end_point_latlng = forms_helper.parse_coordinates(endpoint_latlng)

    if address is not None:
        transaction.address = address

    if merchantid is not None:
        transaction.merchantid = merchantid

    db.session.add(transaction)
    db.session.commit()

    return transaction


def create_sales_transaction(data):
    # Prepare Data
    transaction = Transaction.from_dict(data)
    transaction.start_point_latlng = forms_helper.parse_coordinates(data['start_point_latlng'])
    transaction.end_point_latlng = forms_helper.parse_coordinates(data['end_point_latlng'])

    # Persist
    db.session.add(transaction)
    db.session.commit()

    return transaction


def get_user_sales_transactions(userid, limit=None):
    order_types = ['COVERAGE', '1SS', 'CS3', 'FLEET', 'IIDACS', 'CLIENT VISIT', 'GAS', 'FLIGHT']
    type_list = ["type='" + type + "'" for type in order_types]

    # return Transaction.query \
    #     .filter(Transaction.userid == userid) \
    #     .order_by(desc("(" + ",".join(type_list) + ")"), Transaction.transaction_date) \
    #     .limit(page_size) \
    #     .offset((page_no*page_size) - page_size) \
    #     .all()
    if limit is not None:
        return Transaction.query \
            .filter(Transaction.userid == userid) \
            .order_by(desc("(" + ",".join(type_list) + ")"), Transaction.transaction_date) \
            .limit(limit) \
            .all()

    return Transaction.query \
        .filter(Transaction.userid == userid) \
        .order_by(desc("(" + ",".join(type_list) + ")"), Transaction.transaction_date) \
        .all()


# .order_by("position(type::text in '"+",".join(order_types)+"'") \

# return Transaction.query\
#     .filter(Transaction.userid == userid) \
#     .order_by(desc(Transaction.transaction_date)) \
#     .all()


def add_products_to_branch(branchid, data):
    products = []
    date_released = data['date_released']

    for productid in data['products']:
        datadict = {'branchid': branchid, 'productid': productid, 'date_released': date_released}
        products.append(BranchProduct.from_dict(datadict))

    db.session.add_all(products)
    db.session.commit()


def get_branches_by_filter(q, filter_type):
    limit_num = 2000

    if filter_type == 'name':
        if q:
            return Branch.query.filter(Branch.name.ilike("%" + q.lower() + "%")).limit(limit_num).all()
        else:
            return get_branches_with_limit(500)

    elif filter_type == 'boundary_name':
        boundary = Boundary.query.filter(Boundary.name.ilike("%" + q.lower() + "%")).first()
        if boundary is not None:
            return Branch.query.filter(
                func.ST_DWITHIN(cast(boundary.geometry, Geography), cast(Branch.latlng, Geography), 1)).limit(
                limit_num).all()

    elif filter_type == 'territory':
        territory = Territory.query.filter(Territory.code.ilike("%" + q.lower() + "%")).first()
        if territory is not None:
            return Branch.query.filter(
                func.ST_DWITHIN(cast(territory.geom, Geography), cast(Branch.latlng, Geography), 1)).limit(
                limit_num).all()

    return []


def get_products_for_branches(branch_ids):
    return Branch.query.filter(Branch.id.in_(branch_ids)).all()


def get_branches_within_date_range(start_date, end_date):
    limit_ctr = 1000
    return Branch.query.filter(Branch.operation_started_date.between(start_date, end_date)).limit(limit_ctr).all()


def get_sales_transactions_within_date_range(start_date, end_date, user_id):
    # limit_ctr = 500
    return Transaction.query.filter(Transaction.transaction_date.between(start_date, end_date)).filter(
        Transaction.userid == user_id).order_by(Transaction.transaction_date).all()


def get_sales_transactions_by_date(start_date, user_id):
    limit_ctr = 500
    return Transaction.query.filter(cast_data(Transaction.transaction_date, DATE) == start_date).filter(
        Transaction.userid == user_id).order_by(Transaction.transaction_date).limit(limit_ctr).all()


def get_branches_within_date_range_by_product(start_date, end_date, product):
    query = """SELECT b.*
                FROM branch b
                INNER JOIN branch_product bp
                ON b.id = bp.branchid
                INNER JOIN product p
                ON bp.productid = p.id
                WHERE LOWER(p.name) = '{0}' AND
                bp.date_released BETWEEN '{1}' AND '{2}'
                GROUP BY b.id""".format(product.lower(), start_date, end_date)

    result = db.engine.execute(query).fetchall()

    return [
        Branch.from_dict({
            'id': item[0]
            # 'type': item[3],
            # 'name': item[4],
            # 'adddress': item[5] if item[5] is not None else '',
            # 'latlng': item[6] if item[6] is not None else None,
            # 'status': item[10]
        }) for item in result]


def get_all_branches_count():
    query = 'SELECT COUNT(*) FROM branch'

    result = db.engine.execute(query).fetchone()

    return result[0]


def delete_branch_wihin_boundary(latlngList):
    boundary = forms_helper.parse_area(latlngList)

    result = Branch.query \
        .filter(func.ST_DWITHIN(cast(boundary, Geography), cast(Branch.latlng, Geography), 1)) \
        .all()

    for br in result:
        db.session.delete(br)

    db.session.commit()


def update_sales_transaction_remarks(transactionid, remarks):
    transaction = Transaction.query.get(transactionid)

    transaction.remarks = remarks

    db.session.commit()


def update_sales_transaction_status(transactionid, status):
    transaction = Transaction.query.get(transactionid)

    transaction.status = status

    db.session.commit()


def reverse_geocode(latlng_dict):
    url = 'https://maps.googleapis.com/maps/api/geocode/json??key=' + app.config.get('GOOGLE_MAP_API_KEY')
    url += '&latlng=' + str(latlng_dict['lat']) + ',' + str(latlng_dict['lng'])

    response = requests.get(url, verify=False)

    if response.content:
        content = json.loads(response.content)
        if content['status'] == 'OK' and len(content['results']) > 0:
            result = content['results']
            return result[0]['formatted_address']

    return ''


def geocode(address):
    print "GEOCODING ADDRESS: {0}".format(address)
    url = 'https://maps.googleapis.com/maps/api/geocode/json??key=' + app.config.get('GOOGLE_MAP_API_KEY')
    url += '&address=' + address

    response = requests.get(url, verify=False)

    if response.content:
        content = json.loads(response.content)
        if content['status'] == 'OK' and len(content['results']) > 0:
            result = content['results']
            return result[0]

    return None


def upload_branch_data(file):
    wb = load_workbook(file, read_only=True)

    # grab the active worksheet
    ws = wb.active

    rows = ws.iter_rows()

    columnrow = rows.next()

    for cell in columnrow:
        print "column: {0}".format(cell.value)

    for row in rows:
        time.sleep(0.01)

        data_dict = {}

        raw_name = row[0].value.encode('utf-8')
        fidx = raw_name.find('-')

        data_dict['name'] = raw_name[fidx + 1:].strip() if fidx > -1 else raw_name

        typestr = row[3].value.encode('utf-8').lower()

        if typestr == 'iserv':
            data_dict['type'] = BranchType.GT
        elif typestr == 'mst':
            data_dict['type'] = BranchType.MDC

        lat = row[4].value
        lng = row[5].value
        latlng = None if lat <= 0 or lng <= 0 else {'lat': row[4].value, 'lng': row[5].value}

        if latlng is not None:
            data_dict['latlng'] = forms_helper.parse_coordinates(latlng)
            data_dict['address'] = reverse_geocode(latlng)
        else:
            address_param = data_dict['name'] + ' ' + row[1].value.encode('utf-8')
            result = geocode(address_param)
            if result is not None:
                latlng = result['geometry']['location']
                data_dict['latlng'] = forms_helper.parse_coordinates(latlng)
                data_dict['address'] = result['formatted_address']

        try:
            db.session.add(Branch.from_dict(data_dict))
            db.session.commit()
        except Exception as e:
            # print "ERROR: {0}".format(e)
            db.session.rollback()
            continue

    log.debug("Branch data uploaded")


def upload_branch_sellouts_data(file):
    wb = load_workbook(file, read_only=True)

    # grab the active worksheet
    ws = wb.active

    rows = ws.iter_rows()

    # skip 1 row for columns
    rows.next()

    prev_dict = {}

    for idx, row in enumerate(rows):
        data_dict = {}

        # if idx == 10:
        #     break

        raw_branch_name = row[5].value.encode('utf-8')
        fidx = raw_branch_name.find('-')
        data_dict['branch_name'] = raw_branch_name[fidx + 1:].strip() if fidx > -1 else raw_branch_name

        if row[4].value is not None:
            typestr = row[4].value.encode('utf-8').lower()
            if typestr == 'iserv':
                data_dict['branch_type'] = BranchType.GT
            elif typestr == 'mst':
                data_dict['branch_type'] = BranchType.MDC
            prev_dict['branch_type'] = data_dict['branch_type']
        elif row[4].value is None and idx > 0:
            data_dict['branch_type'] = prev_dict['branch_type']

        # find branch first if exist do not geocode and insert
        branch = Branch.query.filter(Branch.name.ilike('%' + data_dict['branch_name'] + '%')).first()
        if branch is None:
            geocode_result = geocode(data_dict['branch_name'])
            # branch_latlng = None if geocode_result is None else forms_helper.parse_coordinates(geocode_result['geometry']['location'])
            # branch_address = '' if geocode_result is None else geocode_result['formatted_address']
            # branch_dict = {
            #     'name': data_dict['branch_name'],
            #     'type': data_dict['branch_type'],
            #     'latlng': branch_latlng,
            #     'address': branch_address
            # }
            # branch = Branch.from_dict(branch_dict)
            # try:
            #     db.session.add(branch)
            #     db.session.commit()
            #     print "Branch saved. {0}".format(branch_dict)
            # except Exception as e:
            #     db.session.rollback()
            if geocode_result is not None:
                branch_dict = {
                    'name': data_dict['branch_name'],
                    'type': data_dict['branch_type'],
                    'latlng': forms_helper.parse_coordinates(geocode_result['geometry']['location']),
                    'address': geocode_result['formatted_address']
                }
                branch = Branch.from_dict(branch_dict)
                try:
                    db.session.add(branch)
                    db.session.commit()
                    print "Branch saved. {0}".format(branch_dict)
                except Exception as e:
                    # print "ERROR saving branch: {0}".format(e)
                    db.session.rollback()
            else:
                continue

        data_dict['measure_type'] = str(row[6].value)
        data_dict['sellout_amount'] = str(row[7].value)

        if row[0].value is not None:
            data_dict['date'] = str(row[0].value)  # row[0].value.encode('utf-8')
            prev_dict['date'] = data_dict['date']
        elif row[0].value is None and idx > 0:
            data_dict['date'] = prev_dict['date']

        if row[1].value is not None:
            data_dict['brand'] = row[1].value.encode('utf-8')
            prev_dict['brand'] = data_dict['brand']
        elif row[1].value is None and idx > 0:
            data_dict['brand'] = prev_dict['brand']

        if row[2].value is not None:
            data_dict['subbrand'] = row[2].value.encode('utf-8')
            prev_dict['subbrand'] = data_dict['subbrand']
        elif row[2].value is None and idx > 0:
            data_dict['subbrand'] = prev_dict['subbrand']

        if row[3].value is not None:
            data_dict['materialid'] = row[3].value.encode('utf-8')
            prev_dict['materialid'] = data_dict['materialid']
        elif row[3].value is None and idx > 0:
            data_dict['materialid'] = prev_dict['materialid']

        product = Product.query.filter(Product.name.ilike('%' + data_dict['subbrand'] + '%')).first()
        if product is None:
            product_dict = {
                'material_code': data_dict['materialid'],
                'brand': data_dict['brand'],
                'name': data_dict['subbrand']
            }
            product = Product.from_dict(product_dict)
            try:
                db.session.add(product)
                db.session.commit()
                print "Product saved. {0}".format(product_dict)
            except Exception as e:
                # print "ERROR saving product: {0}".format(e)
                db.session.rollback()
        else:  # update product
            product.brand = data_dict['brand']
            product.material_code = data_dict['materialid']
            db.session.commit()
            print "Product Updated. {0}".format(product.to_dict())

        if product:
            branch_product_dict = {
                'branchid': branch.id,
                'productid': product.id
            }
            try:
                db.session.add(BranchProduct.from_dict(branch_product_dict))
                db.session.commit()
                print "BranchProduct saved. {0}".format(branch_product_dict)
            except Exception as e:
                db.session.rollback()
        try:
            sellout_dict = {
                'branchid': branch.id,
                'productid': product.id,
                'sellout_date': data_dict['date'],
                'grossup_amount': data_dict['sellout_amount']
            }
            sellout = Sellout.from_dict(sellout_dict)
            db.session.add(sellout)
            db.session.commit()
            print "Sellout saved. {0}".format(sellout_dict)
        except Exception as e:
            # print "ERROR saving sellout: {0}".format(e)
            db.session.rollback()


def get_branch_sellouts(semester, branch_ids):
    return Sellout.query.filter(
        and_(Sellout.branchid.in_(branch_ids.split(',')), Sellout.sellout_date == semester)).all()
    # return Sellout.query.filter(Sellout.branchid.in_(branch_ids.split(','))).all()


def get_branch_sellouts_by_product(semester, product):
    product = Product.query.filter(Product.name.ilike('%' + product + '%')).first()

    if product is None:
        return []

    return Sellout.query.filter(and_(Sellout.productid == product.id, Sellout.sellout_date == semester)).all()
